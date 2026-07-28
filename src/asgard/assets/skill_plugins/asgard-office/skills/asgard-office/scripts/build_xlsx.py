#!/usr/bin/env python3
"""Sheet spec (YAML/JSON) -> .xlsx.

    asgard skills run asgard-office -- build xlsx SHEET.yaml -o out.xlsx [--values v.json]

A workbook is data plus the rules that produced it. The spec keeps those apart:
`rows` is data, `cells` is formulas, `inputs` marks what a reader may change.
Anything computed goes in as a formula string, never as a Python-computed
number — a sheet that does not recalculate is a screenshot with extra steps.

    title: FY26 plan
    sheets:
      - name: Assumptions
        columns: [{header: Item, width: 34}, {header: Value, width: 14, format: "0.0%"}]
        rows:
          - [Revenue growth, 0.15]
        inputs: [B2]
        notes: {B2: "source: FY25 actuals, board pack p.12"}
      - name: Model
        cells:
          B10: "=SUM(B2:B9)"
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.worksheet.table import Table, TableStyleInfo  # noqa: E402

from officelib import specs, templates  # noqa: E402

# Financial-model colour convention. Blue means "a human typed this"; black means
# "the sheet derived it". Every reviewer of a model reads these before the numbers.
INPUT_FONT = "0000FF"
FORMULA_FONT = "000000"
CROSS_SHEET_FONT = "008000"
CROSS_FILE_FONT = "FF0000"
ASSUMPTION_FILL = "FFFF00"

_CELL = re.compile(r"\A([A-Za-z]{1,3})([1-9][0-9]{0,6})\Z")
_HAIRLINE = Side(style="thin", color="D9DEE5")


def _ref(value: str) -> tuple[int, int]:
    match = _CELL.match(str(value).strip())
    if not match:
        raise ValueError(f"not a cell reference: {value}")
    column, row = match.group(1).upper(), int(match.group(2))
    index = 0
    for char in column:
        index = index * 26 + (ord(char) - 64)
    return row, index


def _column_index(key: object, fallback: int) -> int:
    text = str(key).strip().upper()
    if text.isdigit():
        return int(text)
    if text.isalpha():
        index = 0
        for char in text:
            index = index * 26 + (ord(char) - 64)
        return index
    return fallback


def _load_rows(raw: object, base: Path) -> list[list]:
    if isinstance(raw, str):  # a CSV/TSV path relative to the spec
        path = base / raw if not Path(raw).is_absolute() else Path(raw)
        delimiter = "\t" if path.suffix.lower() in (".tsv", ".tab") else ","
        with path.open(encoding="utf-8-sig", newline="") as handle:
            return [list(row) for row in csv.reader(handle, delimiter=delimiter)]
    if isinstance(raw, list):
        return [list(row) if isinstance(row, (list, tuple)) else [row] for row in raw]
    return []


def _coerce(value):
    """Keep numbers numeric. A number stored as text breaks every formula above it."""
    if isinstance(value, str):
        text = value.strip()
        if re.fullmatch(r"-?\d+", text):
            return int(text)
        if re.fullmatch(r"-?\d*\.\d+", text):
            return float(text)
    return value


class SheetBuilder:
    def __init__(self, theme: specs.Theme, base: Path):
        self.theme = theme
        self.base = base
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)
        self.warnings: list[str] = []
        self.formula_cells = 0

    def add(self, spec: dict) -> None:
        name = str(spec.get("name") or f"Sheet{len(self.workbook.sheetnames) + 1}")[:31]
        sheet = self.workbook.create_sheet(name)
        columns = spec.get("columns") or []
        normalised = [item if isinstance(item, dict) else {"header": str(item)} for item in columns]
        start_row = 1
        title = str(spec.get("title") or "")
        if title:
            cell = sheet.cell(row=1, column=1, value=title)
            cell.font = Font(name=self.theme.font_head, size=self.theme.size_h2, bold=True,
                             color=self.theme.primary)
            start_row = 3

        header_row = start_row if normalised else 0
        if normalised:
            for index, column in enumerate(normalised, start=1):
                cell = sheet.cell(row=header_row, column=index, value=str(column.get("header") or ""))
                cell.font = Font(name=self.theme.font_body, size=self.theme.size_body, bold=True,
                                 color=self.theme.primary)
                cell.fill = PatternFill("solid", fgColor=self.theme.surface)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = Border(bottom=_HAIRLINE)
                width = column.get("width")
                if width:
                    sheet.column_dimensions[get_column_letter(index)].width = float(width)

        rows = _load_rows(spec.get("rows"), self.base)
        offset = header_row + 1 if normalised else start_row
        for row_index, row in enumerate(rows):
            for column_index, value in enumerate(row, start=1):
                cell = sheet.cell(row=offset + row_index, column=column_index, value=_coerce(value))
                cell.font = Font(name=self.theme.font_body, size=self.theme.size_body)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    self.formula_cells += 1

        for reference, value in (spec.get("cells") or {}).items():
            row, column = _ref(reference)
            cell = sheet.cell(row=row, column=column, value=_coerce(value))
            cell.font = Font(name=self.theme.font_body, size=self.theme.size_body, color=FORMULA_FONT)
            if isinstance(value, str) and value.startswith("="):
                self.formula_cells += 1
                if "!" in value and "[" in value:
                    cell.font = Font(name=self.theme.font_body, size=self.theme.size_body, color=CROSS_FILE_FONT)
                elif "!" in value:
                    cell.font = Font(name=self.theme.font_body, size=self.theme.size_body, color=CROSS_SHEET_FONT)

        for reference, fmt in (spec.get("formats") or {}).items():
            if _CELL.match(str(reference)):
                row, column = _ref(reference)
                sheet.cell(row=row, column=column).number_format = str(fmt)
            else:  # a whole column: "B" or "2"
                index = _column_index(reference, 0)
                if not index:
                    self.warnings.append(f"{name}: format target is neither a cell nor a column: {reference}")
                    continue
                for row in range(offset, offset + max(len(rows), 1)):
                    sheet.cell(row=row, column=index).number_format = str(fmt)

        for index, column in enumerate(normalised, start=1):
            fmt = column.get("format")
            if not fmt:
                continue
            for row in range(offset, offset + max(len(rows), 1)):
                sheet.cell(row=row, column=index).number_format = str(fmt)

        for reference in spec.get("inputs") or []:
            row, column = _ref(reference)
            cell = sheet.cell(row=row, column=column)
            cell.font = Font(name=self.theme.font_body, size=self.theme.size_body, color=INPUT_FONT)
            cell.fill = PatternFill("solid", fgColor=ASSUMPTION_FILL)

        for reference, note in (spec.get("notes") or {}).items():
            from openpyxl.comments import Comment

            row, column = _ref(reference)
            sheet.cell(row=row, column=column).comment = Comment(str(note), "Asgard Office")

        freeze = spec.get("freeze")
        if freeze:
            sheet.freeze_panes = str(freeze)
        elif normalised:
            sheet.freeze_panes = f"A{header_row + 1}"

        if spec.get("table") and normalised and rows:
            last = f"{get_column_letter(len(normalised))}{offset + len(rows) - 1}"
            table = Table(displayName=re.sub(r"\W", "", name) or "Data", ref=f"A{header_row}:{last}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
            sheet.add_table(table)

        for index in range(1, max(len(normalised), max((len(row) for row in rows), default=0)) + 1):
            letter = get_column_letter(index)
            if sheet.column_dimensions[letter].width:
                continue
            longest = max(
                (len(str(sheet.cell(row=row, column=index).value or "")) for row in range(1, sheet.max_row + 1)),
                default=10,
            )
            sheet.column_dimensions[letter].width = min(max(10, longest + 2), 60)

        if spec.get("legend"):
            row = sheet.max_row + 2
            sheet.cell(row=row, column=1, value="Legend").font = Font(bold=True, color=self.theme.primary)
            for offset_index, (colour, label) in enumerate(
                (
                    (INPUT_FONT, "blue text on yellow — fill this in"),
                    (FORMULA_FONT, "black — computed, do not overwrite"),
                    (CROSS_SHEET_FONT, "green — links to another sheet"),
                ),
                start=1,
            ):
                cell = sheet.cell(row=row + offset_index, column=1, value=label)
                cell.font = Font(name=self.theme.font_body, size=self.theme.size_caption, color=colour)

    def finish(self, out: Path, meta: specs.DocMeta) -> None:
        if not self.workbook.sheetnames:
            raise ValueError("sheet spec declared no sheets")
        properties = self.workbook.properties
        properties.title = meta.title or properties.title
        properties.creator = meta.author or properties.creator
        out.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(str(out))


def build(
    spec_path: Path,
    out: Path,
    *,
    template_name: str = "",
    values_path: Path | None = None,
    root: Path | None = None,
) -> dict:
    text = spec_path.read_text(encoding="utf-8")
    template = None
    name = template_name
    missing: list[str] = []
    if values_path is not None:
        values = specs.load_data(values_path)
        text, missing = templates.render(text, values)
    raw = json.loads(text) if spec_path.suffix.lower() == ".json" else specs.load_yaml(text)
    if not isinstance(raw, dict):
        raise ValueError("sheet spec must be a mapping with a `sheets` list")
    name = name or str(raw.get("template") or "")
    if name:
        template = templates.resolve(name, root)
        if template.kind != "xlsx":
            raise ValueError(f"template {name!r} builds {template.kind}, not xlsx")
        raw = {**template.defaults, **raw}
        if template.theme:
            raw["theme"] = {**template.theme, **(raw.get("theme") or {})}
    meta = specs.DocMeta.resolve(raw)
    sheets = raw.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        raise ValueError("sheet spec needs a non-empty `sheets` list")
    builder = SheetBuilder(meta.theme, spec_path.resolve().parent)
    for sheet in sheets:
        if not isinstance(sheet, dict):
            raise ValueError("each entry of `sheets` must be a mapping")
        builder.add(sheet)
    builder.finish(out, meta)
    return {
        "output": str(out),
        "template": template.name if template else None,
        "sheets": [str(sheet.get("name") or "") for sheet in sheets],
        "formula_cells": builder.formula_cells,
        "cached_values": False,
        "unresolved_fields": missing,
        "warnings": builder.warnings,
    }


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(prog="build xlsx", description="Sheet spec -> .xlsx")
    parser.add_argument("spec", type=Path)
    parser.add_argument("-o", "--output", type=Path, required=True)
    parser.add_argument("--template", default="")
    parser.add_argument("--values", type=Path, default=None)
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args(argv)
    try:
        report = build(args.spec, args.output, template_name=args.template, values_path=args.values)
    except (ValueError, OSError, KeyError, json.JSONDecodeError) as exc:
        print(f"xlsx build failed: {exc}", file=sys.stderr)
        return 1
    if args.json:
        print(json.dumps(report, ensure_ascii=False, indent=2))
    else:
        print(f"wrote {report['output']}  ({len(report['sheets'])} sheets, {report['formula_cells']} formulas)")
        for warning in report["warnings"]:
            print(f"  warning: {warning}")
        if report["formula_cells"]:
            print("  formulas carry no cached values until a spreadsheet app recalculates — see `verify --lane xlsx`")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
