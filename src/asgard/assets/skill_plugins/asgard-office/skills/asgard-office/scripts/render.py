#!/usr/bin/env python3
"""The external gate: render a document to PDF and page images, or recalculate a workbook.

    asgard skills run asgard-office -- render FILE [-o DIR] [--dpi 150]
    asgard skills run asgard-office -- render --recalc BOOK.xlsx
    asgard skills run asgard-office -- render --probe

Everything else in this skill is pure Python and always available. This is not,
and pretending otherwise would be the actual defect: only a layout engine knows
where a line broke, whether a title wrapped onto two lines and pushed the rule
into the body, or what a formula evaluates to. LibreOffice is that engine.

When it is absent this exits non-zero and names what is missing. A visual check
that silently did not happen is worse than one that never ran.
"""

from __future__ import annotations

import argparse
import json
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

SOFFICE_NAMES = ("soffice", "libreoffice", "soffice.bin")
MAC_PATHS = ("/Applications/LibreOffice.app/Contents/MacOS/soffice",)
WINDOWS_PATHS = (
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
)
INSTALL_HINT = {
    "darwin": "brew install --cask libreoffice",
    "linux": "apt install libreoffice  (or your distribution's package)",
    "win32": "winget install TheDocumentFoundation.LibreOffice",
}


def find_soffice() -> str:
    override = os.environ.get("ASGARD_SOFFICE")
    if override and Path(override).exists():
        return override
    for name in SOFFICE_NAMES:
        found = shutil.which(name)
        if found:
            return found
    for candidate in (*MAC_PATHS, *WINDOWS_PATHS):
        if Path(candidate).exists():
            return candidate
    return ""


def find_rasteriser() -> tuple[str, str]:
    """(tool, kind) for turning a PDF into page images."""
    for name, kind in (("pdftoppm", "poppler"), ("magick", "imagemagick"), ("convert", "imagemagick")):
        found = shutil.which(name)
        if found:
            return found, kind
    return "", ""


def probe() -> dict:
    soffice = find_soffice()
    raster, raster_kind = find_rasteriser()
    return {
        "soffice": soffice,
        "rasteriser": raster,
        "rasteriser_kind": raster_kind,
        "pandoc": shutil.which("pandoc") or "",
        "can_render_pdf": bool(soffice),
        "can_render_images": bool(soffice and raster),
        "can_recalculate": bool(soffice),
        "install_hint": INSTALL_HINT.get(sys.platform, INSTALL_HINT["linux"]),
    }


def _soffice(soffice: str, args: list[str], *, timeout: int) -> subprocess.CompletedProcess:
    """A private profile per call — a bare soffice reuses a user profile and hangs headless."""
    with tempfile.TemporaryDirectory(prefix="asgard-office-profile-") as profile:
        return subprocess.run(
            [
                soffice,
                "--headless",
                "--norestore",
                "--nolockcheck",
                "--nodefault",
                f"-env:UserInstallation=file://{profile}",
                *args,
            ],
            capture_output=True,
            timeout=timeout,
            check=False,
        )


def to_pdf(path: Path, out_dir: Path, *, timeout: int = 180) -> Path:
    soffice = find_soffice()
    if not soffice:
        raise RuntimeError(
            "LibreOffice (soffice) is not installed, so nothing here can lay the document out.\n"
            f"  install: {INSTALL_HINT.get(sys.platform, INSTALL_HINT['linux'])}\n"
            "  or point ASGARD_SOFFICE at the binary.\n"
            "  The static gate still runs without it: asgard skills run asgard-office -- verify FILE"
        )
    out_dir.mkdir(parents=True, exist_ok=True)
    result = _soffice(soffice, ["--convert-to", "pdf", "--outdir", str(out_dir), str(path)], timeout=timeout)
    produced = out_dir / f"{path.stem}.pdf"
    if not produced.is_file():
        raise RuntimeError(
            f"LibreOffice produced no PDF (exit {result.returncode}): "
            f"{result.stderr.decode('utf-8', 'replace')[:400]}"
        )
    return produced


def to_images(pdf: Path, out_dir: Path, *, dpi: int = 150, timeout: int = 180) -> list[Path]:
    tool, kind = find_rasteriser()
    if not tool:
        raise RuntimeError(
            "no PDF rasteriser found — install poppler (`pdftoppm`) or ImageMagick.\n"
            "  The PDF itself was produced; only the page images are missing."
        )
    out_dir.mkdir(parents=True, exist_ok=True)
    prefix = out_dir / f"{pdf.stem}-page"
    for stale in out_dir.glob(f"{pdf.stem}-page*"):
        stale.unlink()
    if kind == "poppler":
        command = [tool, "-jpeg", "-r", str(dpi), str(pdf), str(prefix)]
    else:
        command = [tool, "-density", str(dpi), str(pdf), f"{prefix}-%d.jpg"]
    subprocess.run(command, capture_output=True, timeout=timeout, check=False)
    return sorted(out_dir.glob(f"{pdf.stem}-page*"))


def recalculate(path: Path, *, timeout: int = 180) -> dict:
    """Rewrite a workbook in place with every formula's value cached.

    openpyxl writes formulas as strings and nothing else; until a spreadsheet
    engine evaluates them the file reads back as empty to pandas, to
    `data_only=True`, and to every previewer.
    """
    soffice = find_soffice()
    if not soffice:
        raise RuntimeError(
            "LibreOffice (soffice) is not installed, so formulas cannot be evaluated here.\n"
            f"  install: {INSTALL_HINT.get(sys.platform, INSTALL_HINT['linux'])}\n"
            "  until then the workbook is still correct — it just carries no cached results, "
            "and opening it once in Excel fixes that."
        )
    with tempfile.TemporaryDirectory(prefix="asgard-office-recalc-") as staging:
        result = _soffice(
            soffice,
            ["--convert-to", "xlsx:Calc MS Excel 2007 XML", "--outdir", staging, str(path)],
            timeout=timeout,
        )
        produced = Path(staging) / f"{path.stem}.xlsx"
        if not produced.is_file():
            raise RuntimeError(
                f"recalculation produced nothing (exit {result.returncode}): "
                f"{result.stderr.decode('utf-8', 'replace')[:400]}"
            )
        shutil.copy2(produced, path)
    import openpyxl

    formulas = openpyxl.load_workbook(str(path), data_only=False)
    values = openpyxl.load_workbook(str(path), data_only=True)
    total = cached = 0
    errors: list[str] = []
    for name in formulas.sheetnames:
        for row in formulas[name].iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                total += 1
                result_value = values[name][cell.coordinate].value
                if result_value is not None:
                    cached += 1
                if isinstance(result_value, str) and result_value.startswith("#"):
                    errors.append(f"{name}!{cell.coordinate} = {result_value}")
    return {"file": str(path), "formulas": total, "cached": cached, "errors": errors}


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(prog="render", description="PDF, page images, and workbook recalculation")
    parser.add_argument("file", type=Path, nargs="?")
    parser.add_argument("-o", "--outdir", type=Path, default=None)
    parser.add_argument("--dpi", type=int, default=150)
    parser.add_argument("--pdf-only", action="store_true")
    parser.add_argument("--recalc", action="store_true", help="evaluate an .xlsx and cache the results in place")
    parser.add_argument("--probe", action="store_true", help="report which external tools are available")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args(argv)

    if args.probe:
        report = probe()
        print(json.dumps(report, indent=2) if args.json else "\n".join(f"{k}: {v}" for k, v in report.items()))
        return 0 if report["can_render_pdf"] else 1
    if args.file is None:
        parser.error("a file is required unless --probe is given")
    if not args.file.is_file():
        print(f"file not found: {args.file}", file=sys.stderr)
        return 2

    try:
        if args.recalc:
            report = recalculate(args.file)
            print(json.dumps(report, ensure_ascii=False, indent=2) if args.json else
                  f"recalculated {report['cached']}/{report['formulas']} formulas in {report['file']}")
            for error in report["errors"]:
                print(f"  formula error: {error}")
            return 1 if report["errors"] else 0
        out_dir = args.outdir or (args.file.parent / ".asgard" / ".saga" / args.file.stem)
        pdf = args.file if args.file.suffix.lower() == ".pdf" else to_pdf(args.file, out_dir)
        images: list[Path] = [] if args.pdf_only else to_images(pdf, out_dir, dpi=args.dpi)
        report = {"pdf": str(pdf), "images": [str(image) for image in images]}
        if args.json:
            print(json.dumps(report, ensure_ascii=False, indent=2))
        else:
            print(f"pdf: {pdf}")
            for image in images:
                print(f"page: {image}")
            if images:
                print("\nRead these images — a layout defect is only visible in the render.")
        return 0
    except (RuntimeError, subprocess.TimeoutExpired, OSError) as exc:
        print(f"render gate not satisfied:\n{exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
